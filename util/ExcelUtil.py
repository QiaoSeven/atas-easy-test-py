import json
from collections import OrderedDict

from openpyxl import load_workbook


def get_row_no(sheet, col_name):
    map = {}
    minrow = sheet.min_row
    maxrow = sheet.max_row
    for i in range(minrow, maxrow + 1):
        map[sheet.cell(i, 1).value] = i
    return map[col_name]


class ExcelUtil:
    @staticmethod
    def get_cell(file_fl, col_name, row):
        wb = load_workbook(file_fl)
        sheet = wb.active
        minrow = sheet.min_row
        maxrow = sheet.max_row
        mincol = sheet.min_column
        maxcol = sheet.max_column
        for i in range(minrow, maxrow + 1):
            for j in range(mincol, maxcol + 1):
                cell = sheet.cell(i, j).value
                if cell == col_name:
                    return sheet.cell(i + row, j).value

    @staticmethod
    def get_max_row(file_fl):
        wb = load_workbook(file_fl)
        sheet = wb.active
        return sheet.max_row

    @staticmethod
    def get_data_dict(file_fl, col_name):
        wb = load_workbook(file_fl)
        sheet = wb.active
        num = get_row_no(sheet, col_name)
        mincol = sheet.min_column
        maxcol = sheet.max_column
        map = {}
        for i in range(mincol + 1, maxcol + 1):
            map[sheet.cell(num, i).value] = sheet.cell(num + 1, i).value
        return map

    @staticmethod
    def get_all_data_dict(file_fl):
        wb = load_workbook(file_fl)
        sheet = wb.active
        minrow = sheet.min_row
        maxrow = sheet.max_row
        mincol = sheet.min_column
        maxcol = sheet.max_column
        map = {}
        for i in range(minrow, maxrow + 1):
            temp = OrderedDict()
            name = sheet.cell(i, 1).value
            if name is None:
                i = i + 1
            else:
                for j in range(mincol + 1, maxcol + 1):
                    temp[sheet.cell(i, j).value] = sheet.cell(i + 1, j).value
                map[name] = temp
        return map

    @staticmethod
    def get_all_data_dict_by_name(file_fl, name):
        wb = load_workbook(file_fl)
        sheet = wb.active
        minrow = sheet.min_row
        maxrow = sheet.max_row
        mincol = sheet.min_column
        maxcol = sheet.max_column
        arr = []
        x = 0
        y = 0
        for i in range(minrow, maxrow + 1):
            if sheet.cell(i, 1).value == name:
                x = i
                break
        for j in range(x + 1, maxrow + 1):
            name1 = sheet.cell(j, 1).value
            if name1 is not None:
                y = j - 1
                break

        for i in range(x, y):
            temp = {}
            for j in range(mincol + 1, maxcol + 1):
                temp[sheet.cell(x, j).value] = sheet.cell(i + 1, j).value
            arr.append(temp)
        return arr


if __name__ == '__main__':
    file = '/Users/tao.ding/PycharmProjects/atas-easy-test-py/data/atas-data.xlsx'
    print(ExcelUtil.get_all_data_dict_by_name(file, 'test_login_fail'))
    # print(ExcelUtil.get_multiple_line_data_dict(file)['test_login_fail'])
